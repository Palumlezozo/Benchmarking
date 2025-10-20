#!/usr/bin/env python3
"""
Document Collection Chatbot

An interactive chatbot that queries documents in a collection using OpenAI's file search tool.
Provides structured responses with content, sources, and insights.
"""

import argparse
import asyncio
import json
import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional, List

from openai import OpenAI
from pydantic import BaseModel, Field
from dotenv import load_dotenv
from rich.console import Console
from rich.markdown import Markdown
from rich.panel import Panel
from rich.prompt import Prompt, Confirm

from config import (
    config, add_common_arguments, parse_model_from_args,
    DEFAULT_MODEL, DEFAULT_REASONING_EFFORT, DEFAULT_TEXT_VERBOSITY,
    DEFAULT_EMBEDDING_MODEL
)
from utils import clean_text, clean_markdown_text

# Load environment variables
load_dotenv()

# Constants
STATE_FILE = "data/rag/rag_state.json"
OPENAI_STATE_FILE = "data/rag/openai_state.json"
DEFAULT_COLLECTION = "collection"


class Source(BaseModel):
    """Source information for a response."""
    document: str = Field(description="Name of the source document or website")
    page_number: Optional[int] = Field(None, description="Page number in the document")
    url: Optional[str] = Field(None, description="URL for web sources")
    
    def __str__(self):
        if self.url:
            # Web source with URL
            return f"{self.document} - {self.url}"
        elif self.page_number:
            # Document source with page number
            return f"{self.document} (page {self.page_number})"
        else:
            # Document source without page number
            return self.document


class ChatResponse(BaseModel):
    """Structured response from the chatbot."""
    content: str = Field(description="The main response content in markdown format")
    sources: List[Source] = Field(default_factory=list, description="List of source documents")
    insights: str = Field(description="Additional notes, rationale, or suggested next steps")


class DocumentChatbot:
    """Chatbot for interacting with document collections."""
    
    def __init__(
        self, 
        collections: List[str] = None,
        use_openai_tools: bool = False,
        model: str = DEFAULT_MODEL,
        reasoning_effort: str = DEFAULT_REASONING_EFFORT,
        text_verbosity: str = DEFAULT_TEXT_VERBOSITY,
        embedding_model: str = DEFAULT_EMBEDDING_MODEL
    ):
        """Initialize the chatbot."""
        self.collections = collections or [DEFAULT_COLLECTION]
        self.use_openai_tools = use_openai_tools
        self.model = model
        self.reasoning_effort = reasoning_effort
        self.text_verbosity = text_verbosity
        self.embedding_model = embedding_model
        self.client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        self.console = Console()
        self.vector_store_ids = []  # List of vector store IDs for OpenAI backend
        self.conversation_history = []  # Store conversation history
        self.tool_engine = None  # For my-tools (RAG + web search)
        
        # Load collection state and initialize tools
        if self.use_openai_tools:
            self._load_openai_state()
        else:
            self._initialize_my_tools()
    
    def _load_openai_state(self):
        """Load the OpenAI collection state from the document manager for multiple collections."""
        from pathlib import Path as PathLib
        
        state_file = Path(OPENAI_STATE_FILE)
        
        if not state_file.exists():
            # Check if there are documents in any of the folders
            for collection in self.collections:
                docs_dir = PathLib("data/documents") / collection
                if docs_dir.exists():
                    docs = list(docs_dir.rglob("*"))
                    docs = [d for d in docs if d.is_file() and not d.name.startswith("._")]
                    if docs:
                        self.console.print(f"[yellow]⚠️  OpenAI vector store not initialized but {len(docs)} file(s) found in {docs_dir}[/yellow]")
                        self.console.print(f"[yellow]   Run 'python document_manager.py --openai-tools --collection {collection} --update' to index documents[/yellow]")
            self.console.print(f"[dim]ℹ️  No state file found: {OPENAI_STATE_FILE}[/dim]")
            sys.exit(1)
        
        try:
            with open(state_file, 'r', encoding='utf-8') as f:
                state = json.load(f)
            
            state_collections = state.get("collections", {})
            total_docs = 0
            loaded_collections = []
            
            # Load vector store IDs for each collection
            for collection in self.collections:
                if collection not in state_collections:
                    # Check if there are documents in the folder
                    docs_dir = PathLib("data/documents") / collection
                    if docs_dir.exists():
                        docs = list(docs_dir.rglob("*"))
                        docs = [d for d in docs if d.is_file() and not d.name.startswith("._")]
                        if docs:
                            self.console.print(f"[yellow]⚠️  Collection '{collection}' not indexed but {len(docs)} file(s) found in {docs_dir}[/yellow]")
                            self.console.print(f"[yellow]   Run 'python document_manager.py --openai-tools --collection {collection} --update' to index documents[/yellow]")
                        else:
                            self.console.print(f"[yellow]⚠️  Collection '{collection}' not found[/yellow]")
                    else:
                        self.console.print(f"[yellow]⚠️  Collection '{collection}' not found[/yellow]")
                    continue
                
                collection_state = state_collections[collection]
                vector_store_id = collection_state.get("vector_store_id")
                
                if not vector_store_id:
                    # Check if there are documents in the folder
                    docs_dir = PathLib("data/documents") / collection
                    if docs_dir.exists():
                        docs = list(docs_dir.rglob("*"))
                        docs = [d for d in docs if d.is_file() and not d.name.startswith("._")]
                        if docs:
                            self.console.print(f"[yellow]⚠️  OpenAI vector store is empty for '{collection}' but {len(docs)} file(s) found in {docs_dir}[/yellow]")
                            self.console.print(f"[yellow]   Run 'python document_manager.py --openai-tools --collection {collection} --update' to index documents[/yellow]")
                    continue
                
                # Successfully loaded collection
                self.vector_store_ids.append(vector_store_id)
                doc_count = len(collection_state.get("documents", {}))
                total_docs += doc_count
                loaded_collections.append(f"{collection} ({doc_count} docs)")
            
            if not self.vector_store_ids:
                self.console.print(f"[red]❌ No vector stores found for any of the specified collections[/red]")
                self.console.print(f"[yellow]Available collections: {', '.join(state_collections.keys()) or 'None'}[/yellow]")
                sys.exit(1)
            
            collections_str = ", ".join(loaded_collections)
            self.console.print(f"[green]✅ Connected to collections: {collections_str} (total: {total_docs} documents)[/green]")
            
        except Exception as e:
            self.console.print(f"[red]❌ Error loading state: {e}[/red]")
            sys.exit(1)
    
    def _initialize_my_tools(self):
        """Initialize custom tools (RAG, web search, etc.) for multiple collections."""
        from rag_tool import RAGSearchTool
        from tavily_tool import TavilySearchTool, TavilyNewsSearchTool
        from tool_calling_engine import ToolCallingEngine
        
        try:
            tools = []
            loaded_collections = []
            
            # Create RAG search tools for each collection
            for collection in self.collections:
                try:
                    rag_tool = RAGSearchTool(
                        collection=collection,
                        embedding_model=self.embedding_model
                    )
                    
                    # Check if vector store is empty but documents exist
                    if not rag_tool.rag_store.state.get("index_created", False):
                        docs = rag_tool.rag_store._scan_documents()
                        if docs:
                            self.console.print(f"[yellow]⚠️  RAG vector store is empty for '{collection}' but {len(docs)} document(s) found in {rag_tool.rag_store.documents_dir}[/yellow]")
                            self.console.print(f"[yellow]   Run 'python document_manager.py --collection {collection} --update' to index documents[/yellow]")
                        else:
                            self.console.print(f"[dim]ℹ️  No documents found in collection '{collection}'[/dim]")
                        continue  # Skip this collection
                    
                    # Successfully loaded
                    tools.append(rag_tool)
                    doc_count = rag_tool.rag_store.state.get("total_documents", 0)
                    loaded_collections.append(f"{collection} ({doc_count} docs)")
                    
                except Exception as e:
                    self.console.print(f"[yellow]⚠️  Could not initialize RAG tool for '{collection}': {e}[/yellow]")
            
            # Create Tavily web search tool (uses config defaults)
            try:
                tavily_tool = TavilySearchTool()
                tools.append(tavily_tool)
                
                # Create Tavily news search tool (uses config defaults)
                tavily_news_tool = TavilyNewsSearchTool()
                tools.append(tavily_news_tool)
            except ValueError as e:
                self.console.print(f"[yellow]⚠️  Tavily tools not available: {e}[/yellow]")
                self.console.print("[yellow]   Web search will not be available. Add TAVILY_API_KEY to .env to enable.[/yellow]")
            
            # Ensure we have at least one tool
            if not tools:
                collections_str = "', '".join(self.collections)
                self.console.print(f"[red]❌ No tools available. Please set up at least one of:[/red]")
                self.console.print(f"[red]   1. Index documents for collections: '{collections_str}'[/red]")
                self.console.print(f"[red]      Example: python document_manager.py --collection {self.collections[0]} --update[/red]")
                self.console.print(f"[red]   2. Add TAVILY_API_KEY to .env for web search[/red]")
                sys.exit(1)
            
            # Create tool calling engine with custom tools
            self.tool_engine = ToolCallingEngine(
                tools=tools,
                model=self.model,
                reasoning_effort=self.reasoning_effort,
                text_verbosity=self.text_verbosity
            )
            
            # Build status message
            tool_names = [tool.name() for tool in tools]
            
            # Check if using Azure OpenAI
            azure_base_url = os.getenv("AZURE_OPENAI_BASE_URL")
            provider_info = " (Azure OpenAI)" if azure_base_url else ""
            
            if loaded_collections:
                collections_str = ", ".join(loaded_collections)
                self.console.print(f"[green]✅ Connected to collections: {collections_str}[/green]")
                self.console.print(f"[green]   Available tools: {', '.join(tool_names)}{provider_info}[/green]")
            else:
                self.console.print(f"[green]✅ Tools initialized: {', '.join(tool_names)}{provider_info}[/green]")
            
        except Exception as e:
            self.console.print(f"[red]❌ Error initializing custom tools: {e}[/red]")
            import traceback
            self.console.print(f"[dim]{traceback.format_exc()}[/dim]")
            sys.exit(1)
    
    def query(self, question: str) -> Optional[ChatResponse]:
        """Query the document collection with a question."""
        if self.use_openai_tools:
            return self._query_openai(question)
        else:
            return self._query_my_tools(question)
    
    def _query_openai(self, question: str) -> Optional[ChatResponse]:
        """Query using OpenAI native tools (file_search + web_search) across multiple collections."""
        try:
            # Build collections string
            collections_str = "', '".join(self.collections)
            
            # System prompt for the assistant
            system_prompt = f"""
You are an intelligent document analysis assistant for the collections: '{collections_str}'.

You have access to two tools:
1. **file_search**: Search through the document collection for relevant information
2. **web_search**: Search the web for current information, recent updates, or context not in the documents

IMPORTANT: Your response will be parsed into a structured format with THREE SEPARATE FIELDS:

1. **content** field (string):
   - Contains ONLY the direct answer to the user's question
   - Use markdown formatting (headers, lists, tables, bold, etc.)
   - Do NOT include "SOURCES:" or "INSIGHTS:" headings here
   - Do NOT include reasoning, sources, metadata, suggestions, or recommendations
   - Do NOT include "next steps", "suggested actions", or "you should/could" statements
   - ONLY factual content answering the question
   - Stop after answering - no conclusions beyond the facts

2. **sources** field (list of objects):
   - A list where each source has: "document", "page_number" (optional), and "url" (optional)
   - For documents: {{"document": "Document Name.pdf", "page_number": 123}}
   - For web sources: {{"document": "Source Title", "url": "https://full.url.here"}}
   - Do NOT include source information in the content field

3. **insights** field (string):
   - Additional context, reasoning, and analysis
   - Which tool(s) were used (file_search, web_search, or both) and why
   - Limitations or caveats
   - **SUGGESTED NEXT STEPS** (this is where recommendations go)
   - Any inferences or assumptions made

Guidelines:
- **ALWAYS use file_search** when the user asks about documents, collections, or content in the documents
- **ALWAYS use file_search FIRST** before answering questions about the collection
- If asked about "the document" or "documents in the collection", use file_search to find and retrieve content
- Use web_search only for current events, recent updates, or general knowledge not in documents
- Be specific and accurate
- Do not include citation markers like [cite:X] anywhere
- Keep content focused on facts only
- Put ALL reasoning, context, and suggestions in insights, NOT in content
- Consider conversation history for follow-up questions

Example structure (what you should produce):
{{
  "content": "# Answer\n\n• Fact 1\n• Fact 2\n• Fact 3",
  "sources": [
    {{"document": "Report.pdf", "page_number": 5}},
    {{"document": "Website Title", "url": "https://example.com"}}
  ],
  "insights": "Used file_search and web_search tools. Data from 2024 report. Suggested next steps: Review the detailed breakdown in section 5 for more context."
}}
"""
            
            # Build the message list with conversation history
            messages = [{"role": "system", "content": system_prompt}]
            
            # Add conversation history
            messages.extend(self.conversation_history)
            
            # Add current question
            messages.append({"role": "user", "content": question})
            
            # Use responses API with file_search and web_search tools
            response = self.client.responses.parse(
                model=self.model,
                input=messages,
                text_format=ChatResponse,  # Use Pydantic model directly
                tools=[
                    {
                        "type": "file_search",
                        "vector_store_ids": self.vector_store_ids  # Support multiple collections
                    },
                    {
                        "type": "web_search"
                    }
                ],
                tool_choice="auto",
                reasoning={"effort": self.reasoning_effort},
                text={"verbosity": self.text_verbosity}
            )
            
            # Extract the parsed response - get the LAST assistant message (after tool execution)
            chat_response = None
            if hasattr(response, 'output') and response.output:
                # Collect all assistant messages (don't break on first one)
                all_responses = []
                for item in response.output:
                    # Check if item has type 'message' and role 'assistant'
                    if hasattr(item, 'type') and item.type == 'message' and hasattr(item, 'role') and item.role == 'assistant':
                        # Check if content exists
                        if hasattr(item, 'content') and item.content:
                            # Content is a list, iterate through it
                            if isinstance(item.content, list):
                                for content_item in item.content:
                                    # Look for output_text type with parsed content
                                    if hasattr(content_item, 'type') and content_item.type == 'output_text':
                                        if hasattr(content_item, 'text_parsed') and content_item.text_parsed:
                                            all_responses.append(content_item.text_parsed)
                                            break
                                        elif hasattr(content_item, 'parsed') and content_item.parsed:
                                            all_responses.append(content_item.parsed)
                                            break
                                        elif hasattr(content_item, 'text') and isinstance(content_item.text, ChatResponse):
                                            all_responses.append(content_item.text)
                                            break
                            # If content itself is the parsed object
                            elif isinstance(item.content, ChatResponse):
                                all_responses.append(item.content)
                
                # Use the LAST response (final answer after tool execution)
                if all_responses:
                    chat_response = all_responses[-1]
            
            if not chat_response:
                self.console.print("[red]❌ No response content found[/red]")
                return None
            
            # Get raw text for conversation history
            response_text = json.dumps({
                "content": chat_response.content,
                "sources": [{"document": s.document, "page_number": s.page_number} for s in chat_response.sources],
                "insights": chat_response.insights
            })
            
            # Add to conversation history
            self.conversation_history.append({"role": "user", "content": question})
            self.conversation_history.append({"role": "assistant", "content": response_text})
            
            return chat_response
            
        except Exception as e:
            self.console.print(f"[red]❌ Error querying documents: {e}[/red]")
            return None
    
    def _query_my_tools(self, question: str) -> Optional[ChatResponse]:
        """Query using custom tools (RAG, web search, etc.) - handles async internally."""
        try:
            # Build collections string
            collections_str = "', '".join(self.collections)
            
            # System prompt for the assistant
            system_prompt = f"""
You are an intelligent document analysis assistant for the collections: '{collections_str}'.

You have access to multiple tools:
- **search_documents**: Search through the document collection for relevant information
- **web_search**: Search the web for current information, recent updates, or general knowledge
- **news_search**: Search for recent news articles and current events

IMPORTANT: Your response will be parsed into a structured format with THREE SEPARATE FIELDS:

1. **content** field (string):
   - Contains ONLY the direct answer to the user's question
   - Use markdown formatting (headers, lists, tables, bold, etc.)
   - Do NOT include "SOURCES:" or "INSIGHTS:" headings here
   - Do NOT include reasoning, sources, metadata, suggestions, or recommendations
   - Do NOT include "next steps", "suggested actions", or "you should/could" statements
   - ONLY factual content answering the question
   - Stop after answering - no conclusions beyond the facts

2. **sources** field (list of objects):
   - A list where each source has: "document", "page_number" (optional), and "url" (optional)
   - For documents: {{"document": "Document Name.pdf", "page_number": 123}}
   - For web sources: {{"document": "Source Title", "url": "https://full.url.here"}}
   - Extract these from tool results
   - Do NOT include source information in the content field

3. **insights** field (string):
   - Additional context, reasoning, and analysis
   - Which tool(s) were used and why
   - Limitations or caveats
   - **SUGGESTED NEXT STEPS** (this is where recommendations go)
   - Any inferences or assumptions made

Guidelines:
- Use the appropriate tools based on the question type
- Base your answers strictly on the tool results
- Extract and structure source information correctly
- Do not include citation markers like [cite:X] anywhere
- Keep content focused on facts only
- Put ALL reasoning, context, and suggestions in insights, NOT in content
- Consider conversation history for follow-up questions

Example structure (what you should produce):
{{
  "content": "# Answer\n\n• Fact 1\n• Fact 2\n• Fact 3",
  "sources": [
    {{"document": "Report.pdf", "page_number": 5}},
    {{"document": "Website Title", "url": "https://example.com"}}
  ],
  "insights": "Used search_documents and web_search tools. Data from 2024 report. Suggested next steps: Review the detailed breakdown in section 5 for more context."
}}
"""
            
            # Use tool calling engine (now async, so we need to run it)
            chat_response = asyncio.run(
                self.tool_engine.query(
                    question=question,
                    conversation_history=self.conversation_history,
                    response_format=ChatResponse,
                    system_prompt=system_prompt
                )
            )
            
            if not chat_response:
                self.console.print("[red]❌ No response from tool engine[/red]")
                return None
            
            # Get raw text for conversation history
            response_text = json.dumps({
                "content": chat_response.content,
                "sources": [{"document": s.document, "page_number": s.page_number} for s in chat_response.sources],
                "insights": chat_response.insights
            })
            
            # Add to conversation history
            self.conversation_history.append({"role": "user", "content": question})
            self.conversation_history.append({"role": "assistant", "content": response_text})
            
            return chat_response
            
        except Exception as e:
            self.console.print(f"[red]❌ Error querying with custom tools: {e}[/red]")
            import traceback
            self.console.print(f"[dim]{traceback.format_exc()}[/dim]")
            return None
    
    def clear_history(self):
        """Clear the conversation history."""
        self.conversation_history = []
        self.console.print("[yellow]🔄 Conversation history cleared[/yellow]")
    
    def get_history_summary(self) -> str:
        """Get a summary of the conversation history."""
        if not self.conversation_history:
            return "No conversation history"
        
        # Count user messages (every other message starting from 0)
        user_message_count = len([m for m in self.conversation_history if m["role"] == "user"])
        return f"{user_message_count} message(s) in history"
    
    def _detect_markdown_tables(self, content: str) -> List[List[List[str]]]:
        """
        Detect and extract all markdown tables from the content.
        
        Returns:
            List of tables, where each table is a list of rows, and each row is a list of cell values.
            Returns empty list if no markdown tables found.
        """
        lines = content.split('\n')
        all_tables = []
        current_table = []
        in_markdown_table = False
        
        for i, line in enumerate(lines):
            stripped = line.strip()
            
            # Check if line is a markdown table row (starts and ends with |)
            if stripped.startswith('|') and stripped.endswith('|'):
                # Skip separator rows (like |---|---|)
                if all(c in '|-: ' for c in stripped):
                    continue
                
                # Extract cells from markdown table
                cells = [cell.strip() for cell in stripped.split('|')[1:-1]]
                if cells and len(cells) >= 2:  # At least 2 columns
                    current_table.append(cells)
                    in_markdown_table = True
            elif in_markdown_table:
                # End of current markdown table
                if len(current_table) >= 2:  # At least header + one data row
                    all_tables.append(current_table)
                current_table = []
                in_markdown_table = False
        
        # Don't forget the last table if we ended while still in it
        if in_markdown_table and len(current_table) >= 2:
            all_tables.append(current_table)
        
        return all_tables
    
    def _save_to_excel(self, table_data: List[List[str]], filename: Optional[str] = None) -> bool:
        """Save tabular data to an Excel file with formatting."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            if not filename:
                # Auto-generate filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.xlsx"
            
            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Export"
            
            # Write data to worksheet
            for row_idx, row_data in enumerate(table_data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    # Format header row (first row)
                    if row_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    # Add borders
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to file
            filepath = Path(filename)
            wb.save(filepath)
            
            self.console.print(f"[green]✅ Excel file saved to: {filepath.absolute()}[/green]")
            return True
        except Exception as e:
            self.console.print(f"[red]❌ Error saving Excel file: {e}[/red]")
            import traceback
            self.console.print(f"[dim]{traceback.format_exc()}[/dim]")
            return False
    
    def display_response(self, response: ChatResponse, interactive: bool = True):
        """Display a formatted response."""
        self.console.print()
        
        # Clean citation artifacts from content while preserving markdown structure
        cleaned_content = clean_markdown_text(response.content)
        
        # Display content as markdown
        self.console.print(Panel.fit(
            Markdown(cleaned_content),
            title="[bold blue]Response[/bold blue]",
            border_style="blue"
        ))
        
        # Display sources (sorted by document name, then page number)
        if response.sources:
            # Sort sources: by document name first, then by page number (None last)
            sorted_sources = sorted(
                response.sources,
                key=lambda s: (s.document, s.page_number if s.page_number is not None else float('inf'))
            )
            # Clean citation artifacts from source names
            sources_text = "\n".join([f"• {clean_text(str(source))}" for source in sorted_sources])
            self.console.print(Panel.fit(
                sources_text,
                title="[bold green]Sources[/bold green]",
                border_style="green"
            ))
        
        # Clean citation artifacts from insights
        cleaned_insights = clean_text(response.insights)
        
        # Display insights
        self.console.print(Panel.fit(
            cleaned_insights,
            title="[bold yellow]Insights[/bold yellow]",
            border_style="yellow"
        ))
        
        # Check if content contains markdown tables and offer to save to Excel (use cleaned content)
        markdown_tables = self._detect_markdown_tables(cleaned_content)
        if markdown_tables and interactive:
            self.console.print()
            table_count = len(markdown_tables)
            plural = "table" if table_count == 1 else "tables"
            if Confirm.ask(f"📊 [cyan]{table_count} markdown {plural} detected. Would you like to save to Excel?[/cyan]", default=True):
                # Save each table to a separate file
                for i, table_data in enumerate(markdown_tables, 1):
                    if table_count > 1:
                        self.console.print(f"[cyan]Table {i} of {table_count}:[/cyan]")
                        filename = Prompt.ask(f"💾 [cyan]Enter filename for table {i} (press Enter for auto-generated)[/cyan]", default="")
                    else:
                        filename = Prompt.ask("💾 [cyan]Enter filename (press Enter for auto-generated)[/cyan]", default="")
                    
                    if not filename:
                        # Auto-generate filename with table number if multiple tables
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        if table_count > 1:
                            filename = f"export_{timestamp}_table{i}.xlsx"
                        else:
                            filename = f"export_{timestamp}.xlsx"
                    
                    self._save_to_excel(table_data, filename)
        elif markdown_tables and not interactive:
            # Auto-save in non-interactive mode - each table to a separate file
            table_count = len(markdown_tables)
            plural = "table" if table_count == 1 else "tables"
            self.console.print(f"[cyan]📊 {table_count} markdown {plural} detected - auto-saving to Excel...[/cyan]")
            for i, table_data in enumerate(markdown_tables, 1):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                if table_count > 1:
                    filename = f"export_{timestamp}_table{i}.xlsx"
                else:
                    filename = f"export_{timestamp}.xlsx"
                self._save_to_excel(table_data, filename)
        
        self.console.print()
    
    def interactive_mode(self):
        """Run the chatbot in interactive mode."""
        collections_str = "', '".join(self.collections)
        tools_name = "OpenAI (file_search + web_search)" if self.use_openai_tools else "RAG + Web Search (my-tools)"
        self.console.print(Panel.fit(
            f"[bold]Document Collection Chatbot[/bold]\n"
            f"Collections: [cyan]'{collections_str}'[/cyan]\n"
            f"Tools: [cyan]{tools_name}[/cyan]\n"
            f"Model: [cyan]{self.model}[/cyan]\n\n"
            f"💬 Conversation history is maintained across questions\n"
            f"Commands:\n"
            f"  • Type your questions normally\n"
            f"  • [yellow]'clear'[/yellow] - Clear conversation history\n"
            f"  • [yellow]'history'[/yellow] - Show history summary\n"
            f"  • [yellow]'exit'[/yellow] or [yellow]'quit'[/yellow] - Exit chatbot",
            border_style="magenta"
        ))
        
        while True:
            try:
                # Show history summary in prompt
                history_info = f" [{self.get_history_summary()}]" if self.conversation_history else ""
                question = Prompt.ask(f"\n[bold cyan]You{history_info}[/bold cyan]")
                
                if question.lower() in ['exit', 'quit', 'q']:
                    self.console.print("[yellow]Goodbye![/yellow]")
                    break
                
                if question.lower() == 'clear':
                    self.clear_history()
                    continue
                
                if question.lower() == 'history':
                    self.console.print(f"[cyan]{self.get_history_summary()}[/cyan]")
                    if self.conversation_history:
                        self.console.print("\n[dim]Recent messages:[/dim]")
                        for i, msg in enumerate(self.conversation_history[-6:], 1):  # Show last 3 exchanges
                            role = "You" if msg["role"] == "user" else "Assistant"
                            preview = msg["content"][:100] + "..." if len(msg["content"]) > 100 else msg["content"]
                            self.console.print(f"[dim]{role}: {preview}[/dim]")
                    continue
                
                if not question.strip():
                    continue
                
                # Query the documents
                with self.console.status("[bold green]Thinking...[/bold green]"):
                    response = self.query(question)
                
                if response:
                    self.display_response(response)
                
            except KeyboardInterrupt:
                self.console.print("\n[yellow]Goodbye![/yellow]")
                break
            except Exception as e:
                self.console.print(f"[red]❌ Error: {e}[/red]")
    
    def single_query_mode(self, question: str):
        """Run a single query and display the result."""
        with self.console.status("[bold green]Thinking...[/bold green]"):
            response = self.query(question)
        
        if response:
            self.display_response(response, interactive=False)
        else:
            sys.exit(1)


def main():
    """Main function for the Document Chatbot."""
    parser = argparse.ArgumentParser(
        description="Document Collection Chatbot - Query documents using natural language (uses RAG + web search by default)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode with default collection (RAG + web search)
  python chatbot.py
  
  # Interactive mode with OpenAI native tools
  python chatbot.py --openai-tools --collection "my_docs"
  python chatbot.py --openai-tools --collections "my_docs"  # Both work!
  
  # Query multiple collections at once
  python chatbot.py --collection "collection1,collection2,collection3"
  python chatbot.py --collections "collection1,collection2,collection3"
  
  # Single query mode (uses RAG + web search by default)
  python chatbot.py --query "What are the main features?"
  
  # Query across multiple collections with OpenAI tools
  python chatbot.py --openai-tools --collection "docs1,docs2" --query "What are the trends?"
  
  # With specific model configuration
  python chatbot.py --collection "my_docs" --model gpt-5 --reasoning-effort high
  
  # Light mode (fast and economical)
  python chatbot.py --light --query "scope 3 emissions"
  
  # Show API call logs for debugging
  python chatbot.py --info --query "What are the emissions?"
        """
    )
    
    parser.add_argument(
        "--collection", "--collections",
        dest="collection",
        default=DEFAULT_COLLECTION,
        help=f"Collection name(s) to query. Use comma-separated values for multiple collections (default: '{DEFAULT_COLLECTION}')"
    )
    parser.add_argument(
        "--openai-tools",
        action="store_true",
        help="Use OpenAI native tools (file_search + web_search) instead of RAG + web search (default: RAG + web search)"
    )
    parser.add_argument(
        "--query",
        help="Single query to execute (non-interactive mode)"
    )
    parser.add_argument(
        "--light",
        action="store_true",
        help="Use lightweight model configuration (gpt-5-nano, low reasoning, low verbosity)"
    )
    parser.add_argument(
        "--info",
        action="store_true",
        help="Show detailed API call information and debug logs"
    )
    
    # Add common configuration arguments
    add_common_arguments(parser, include_concurrent=False, include_timeout=False, include_similarity=False, include_embedding_model=True)
    
    args = parser.parse_args()
    
    # Configure logging based on --info flag
    if args.info:
        # Show INFO level logs (API calls, HTTP requests, etc.)
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
    else:
        # Only show WARNING and above (suppress INFO and DEBUG)
        logging.basicConfig(
            level=logging.WARNING,
            format='%(levelname)s: %(message)s'
        )
        # Suppress specific noisy loggers
        logging.getLogger("httpx").setLevel(logging.WARNING)
        logging.getLogger("openai").setLevel(logging.WARNING)
        logging.getLogger("chromadb").setLevel(logging.WARNING)
        logging.getLogger("llama_index").setLevel(logging.WARNING)
    
    # Parse model configuration from arguments
    try:
        # If --light is specified, override with lightweight configuration
        if args.light:
            model = "gpt-5-nano"
            reasoning_effort = "low"
            text_verbosity = "low"
        else:
            model, reasoning_effort, text_verbosity = parse_model_from_args(args)
    except Exception as e:
        print(f"❌ Error parsing model configuration: {e}")
        sys.exit(1)
    
    # Parse collections (support comma-separated values)
    collections = [c.strip() for c in args.collection.split(',')]
    
    # Initialize the chatbot
    try:
        chatbot = DocumentChatbot(
            collections=collections,
            use_openai_tools=args.openai_tools,
            model=model,
            reasoning_effort=reasoning_effort,
            text_verbosity=text_verbosity,
            embedding_model=args.embedding_model
        )
    except Exception as e:
        print(f"❌ Error initializing chatbot: {e}")
        sys.exit(1)
    
    # Run in appropriate mode
    if args.query:
        chatbot.single_query_mode(args.query)
    else:
        chatbot.interactive_mode()


if __name__ == "__main__":
    main()

