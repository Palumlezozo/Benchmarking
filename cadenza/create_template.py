#!/usr/bin/env python3
"""
Create a template Excel file for statement verification.

This script creates an example Excel file showing the expected format
for the verify_statements.py script.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_template(output_file: str = "template.xlsx"):
    """Create an example template file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Statements"
    
    # Headers
    headers = ["ID", "contexts", "category", "priority"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Example data
    data = [
        [1, "The company achieved 100% renewable energy usage in 2024", "Energy", "High"],
        [2, "Carbon emissions were reduced by 50% compared to baseline year", "Emissions", "High"],
        [3, "All suppliers are ISO 14001 certified", "Supply Chain", "Medium"],
        [4, "Employee satisfaction score increased to 85%", "Social", "Medium"],
        [5, "Zero waste to landfill was achieved across all facilities", "Waste", "High"],
        [6, "Water consumption decreased by 30%", "Water", "Medium"],
        [7, "100% of packaging materials are recyclable", "Materials", "Low"],
        [8, "Gender pay gap was eliminated in all departments", "Social", "High"],
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    wb.save(output_file)
    print(f"Template created: {output_file}")
    print(f"\nTo use this template:")
    print(f"  python cadanza/verify_statements.py --input {output_file} --collection YOUR_COLLECTION")


if __name__ == "__main__":
    create_template()

