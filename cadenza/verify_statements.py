#!/usr/bin/env python3
"""
IRO Verification Script

This script reads an XLSX file with 'questions' and 'contexts' columns, extracts
IROs (Impacts, Risks, Opportunities) and their stated facts, verifies the facts
against a specified document collection using RAG, and outputs a new XLSX with
verification results including:
- IRO Statement: Clear, concise description of the Impact/Risk/Opportunity
- Stated Facts: Facts from the context supporting the IRO
- Facts Verified: Whether documents support the stated facts (yes/no/partial)
- Evidence Description: What was found in the documents
- Sources Consulted: Document names and page numbers
"""

import argparse
import asyncio
import logging
import os
import sys
import yaml
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any
from datetime import datetime

from pydantic import BaseModel, Field
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from rich.console import Console
from rich.progress import Progress, TaskID, SpinnerColumn, BarColumn, TextColumn, TimeElapsedColumn

# Add both parent directory and llm_v2 directory to path
# This allows importing llm_v2 as a package AND allows modules within llm_v2 to import each other
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(project_root / "llm_v2"))

from llm_v2.rag_tool import RAGSearchTool
from llm_v2.tavily_tool import TavilySearchTool
from llm_v2.tool_calling_engine import ToolCallingEngine
from llm_v2.config import (
    DEFAULT_MODEL, DEFAULT_REASONING_EFFORT, DEFAULT_TEXT_VERBOSITY,
    DEFAULT_EMBEDDING_MODEL
)
from llm_v2.utils import clean_text

# Load environment variables
load_dotenv()

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configuration
MAX_CONCURRENT_VERIFICATIONS = 5  # Process up to 5 statements in parallel
BATCH_SIZE = 10  # Save checkpoint after every 10 IROs
MAX_RETRIES = 3  # Retry failed IROs up to 3 times


class VerificationResult(BaseModel):
    """Result of verifying an IRO (Impact/Risk/Opportunity) against the document collection."""
    iro_statement: str = Field(
        description="The Impact, Risk, or Opportunity being evaluated, stated clearly and concisely in one sentence. MUST use modal language (may, might, could, would, should) or hedging language (likely, potentially, possibly, appears to). This expresses a POSSIBILITY, not a fact."
    )
    stated_facts: str = Field(
        description="The concrete factual claims from the context that support the IRO evaluation, listed as bullet points or numbered list. EXCLUDE the IRO itself - only include factual claims about reality, not the possibility being evaluated."
    )
    facts_verified: str = Field(
        description="Whether the stated facts are supported by documents: 'yes', 'no', or 'partial'"
    )
    evidence_description: str = Field(
        description="Description of the evidence found (from documents or web) that supports or refutes the stated facts. Clearly indicate which facts were verified by which source type."
    )
    sources_consulted: str = Field(
        description="List of ALL sources consulted. Format: 'Document Name (page X)' for documents and 'Web: [URL]' for web sources. Example: 'Annual Report 2024 (page 15), Web: https://example.com'"
    )
    new_question: str = Field(
        description="A reformulated question based on the IRO statement. Should read neutrally without implying that an effect already exists. Frame as an open inquiry. Use VARIED question formulations based on IRO type (Risk/Opportunity/Positive Impact/Negative Impact). Rotate between patterns like: 'To what extent...', 'How [adjective]...', 'What level of...', 'What degree of...'. Use varied adjectives: concerning, threatening, severe, damaging, beneficial, valuable, significant, critical, meaningful."
    )
    new_context: str = Field(
        description="A reformulated context based on the VERIFIED facts only. Present the verified facts in 2-4 clear, factual sentences. Only include facts that were actually verified (yes or partial verification). DO NOT include any source mentions, citations, document names, page numbers, or URLs - only the factual content itself."
    )


class IROCheckpoint(BaseModel):
    """Checkpoint data for an IRO verification."""
    row_idx: int
    question: str
    context: str
    status: str  # 'pending', 'completed', 'error'
    retry_count: int = 0
    result: Optional[VerificationResult] = None
    error_message: Optional[str] = None
    last_updated: str = Field(default_factory=lambda: datetime.now().isoformat())


class StatementVerifier:
    """Verifies IROs (Impacts, Risks, Opportunities) against a document collection using RAG."""
    
    def __init__(
        self,
        collection: str,
        model: str = DEFAULT_MODEL,
        reasoning_effort: str = DEFAULT_REASONING_EFFORT,
        text_verbosity: str = DEFAULT_TEXT_VERBOSITY,
        embedding_model: str = DEFAULT_EMBEDDING_MODEL
    ):
        """
        Initialize the statement verifier.
        
        Args:
            collection: Name of the document collection to search
            model: OpenAI model to use
            reasoning_effort: Reasoning effort level
            text_verbosity: Text verbosity level
            embedding_model: Embedding model for RAG
        """
        self.collection = collection
        self.model = model
        self.reasoning_effort = reasoning_effort
        self.text_verbosity = text_verbosity
        self.embedding_model = embedding_model
        self.console = Console()
        
        # Track recent question patterns to enforce variation
        self.recent_question_patterns = []
        self.pattern_cycle_index = 0
        self.question_starters = [
            "To what extent",
            "How",
            "What level of",
            "What degree of",
            "What is the magnitude of",
            "How significant is",
            "How substantial",
            "To what degree"
        ]
        
        # Check Azure OpenAI credentials
        azure_base_url = os.getenv("AZURE_OPENAI_BASE_URL")
        azure_api_key = os.getenv("AZURE_OPENAI_API_KEY")
        
        if not azure_base_url or not azure_api_key:
            raise ValueError(
                "Azure OpenAI credentials not found. Please set AZURE_OPENAI_BASE_URL "
                "and AZURE_OPENAI_API_KEY in your .env file."
            )
        
        # Initialize RAG tool
        logger.info(f"Initializing RAG tool for collection: {collection}")
        rag_tool = RAGSearchTool(
            collection=collection,
            embedding_model=embedding_model
        )
        
        # Initialize Tavily web search tool (as fallback)
        tools = [rag_tool]
        tavily_api_key = os.getenv("TAVILY_API_KEY")
        if tavily_api_key:
            logger.info("Initializing Tavily web search tool for fallback verification")
            web_search_tool = TavilySearchTool(
                max_results=3,  # Keep it light
                include_answer=False,
                include_raw_content=False
            )
            tools.append(web_search_tool)
        else:
            logger.warning("TAVILY_API_KEY not found - web search fallback will not be available")
        
        # Initialize tool calling engine with RAG and web search tools
        logger.info(f"Initializing tool calling engine with {len(tools)} tools")
        self.engine = ToolCallingEngine(
            tools=tools,
            model=model,
            reasoning_effort=reasoning_effort,
            text_verbosity=text_verbosity
        )
    
    def _get_pattern_enforcement_instruction(self) -> str:
        """
        Generate instruction to enforce question pattern variation.
        Uses a dimension-based approach to create tailored magnitude questions.
        """
        instruction = """
**CRITICAL - QUESTION FORMULATION REQUIREMENTS:**

You are generating a survey question for a materiality assessment (1-5 magnitude scale).

STEP 1: Infer the most relevant materiality dimension for this IRO:
- "Financial / Business Impact" → revenue, cost, profit, efficiency, fines, liabilities, reformulation, operations, strategic upside/downside
- "Stakeholder / Societal Impact" → employees' well-being, safety, customer/user transparency, equal opportunity, community effects, supplier practices, rights
- "Governance / Compliance Impact" → codes of conduct, whistleblowing, integrity culture, anti-corruption, policy controls
- "Financial & Reputational Risk" → cyber/data/privacy incidents, large trust shocks that create financial exposure and brand damage
- "Compliance / Financial Risk" → non-compliance with laws/standards causing financial exposure, penalties, loss of trust

STEP 2: Generate ONE clear, concise question tailored to that dimension:
- DO NOT include the scale in the question (respondents know it's 1-5)
- DO NOT use "To what extent" as the question starter
- DO NOT add explanations, notes, or extra fields - ONLY the question
- Keep each question under 28 words when possible
- Keep it specific to the IRO context
- Use natural, varied wording - ensure stylistic variation to avoid monotonous repetition
- Make the question GENERIC (not company-specific) - use "the organization" or "the company" instead of specific company names

Phrasing templates by dimension (vary naturally):

**Financial / Business Impact:**
- "How significant could the financial benefit be if [X]?"
- "How material could the business impact be from [X]?"
- "What size impact could [X] have on revenue, cost, or margin?"
- "How substantial is the potential operational or strategic value from [X]?"

**Stakeholder / Societal Impact:**
- "How strongly might [X] affect [stakeholder] [well-being/safety/fairness/informed choice]?"
- "How meaningful is the stakeholder benefit (or harm) from [X]?"
- "How important could [X] be for customer/employee trust or outcomes?"
- "What degree of impact could [X] have on [stakeholder group]?"

**Governance / Compliance Impact:**
- "How effectively might [X] strengthen integrity and reduce misconduct risk?"
- "How much could [X] improve governance controls across the value chain?"
- "How valuable could [X] be for building an ethical culture?"

**Financial & Reputational Risk:**
- "How severe would the financial and reputational consequences be if [event] occurred?"
- "How large could the trust and liability impact be from [event]?"
- "What magnitude of financial loss and brand damage could result from [event]?"

**Compliance / Financial Risk:**
- "How substantial could the financial exposure be from non-compliance with [rule/standard]?"
- "What is the potential cost and brand impact from failing to meet [requirement]?"
- "How serious could the penalties and trust loss be from [compliance failure]?"

REMEMBER: 
- Pick the MOST relevant dimension
- Create ONE specific question (no explanations, notes, or extra fields)
- Keep it under 28 words when possible
- Avoid "To what extent"
- Keep it generic (not company-specific)
- Natural, varied phrasing within the dimension's style to avoid monotonous repetition
"""
        return instruction
    
    def _track_question_pattern(self, question: str):
        """Extract and track the pattern used in a generated question."""
        # Extract first 2-4 words as the pattern
        words = question.split()[:4]
        pattern = ' '.join(words[:3]) if len(words) >= 3 else ' '.join(words)
        
        self.recent_question_patterns.append(pattern)
        if len(self.recent_question_patterns) > 5:
            self.recent_question_patterns.pop(0)
        
        # Advance cycle index for next question
        self.pattern_cycle_index += 1
        
        logger.debug(f"Tracked question pattern: {pattern}")
    
    async def verify_iro(self, question: str, context: str) -> VerificationResult:
        """
        Verify an IRO (Impact/Risk/Opportunity) by extracting the IRO statement,
        identifying stated facts, and verifying those facts against documents.
        
        Args:
            question: The question that evaluates the IRO
            context: The context containing facts that support the IRO evaluation
            
        Returns:
            VerificationResult with IRO statement, stated facts, and verification
        """
        logger.info(f"Verifying IRO - Question: {question[:100]}...")
        
        # System prompt for the tool calling engine
        # Check if web search is available
        has_web_search = os.getenv("TAVILY_API_KEY") is not None
        
        tools_description = f"- search_documents_{self.collection}: Search the '{self.collection}' collection for relevant information"
        if has_web_search:
            tools_description += "\n- web_search: Search the web for information not found in documents (use as fallback)"
        
        system_prompt = f"""You are an expert analyst evaluating IROs (Impacts, Risks, and Opportunities) in corporate sustainability reporting.

Available tools:
{tools_description}

YOUR TASK:

You will receive:
1. A QUESTION that evaluates a potential Impact, Risk, or Opportunity (IRO) - this is a POSSIBILITY
2. A CONTEXT with factual claims that are meant to support the IRO evaluation - these are FACTS

Your job is to:
1. **Extract the IRO**: From the question, identify the specific Impact, Risk, or Opportunity being evaluated (1 clear, concise sentence about the POSSIBILITY). **CRITICAL: The IRO statement MUST use modal or hedging language.**
2. **Identify Stated Facts**: From the context, list ONLY the factual claims - EXCLUDE the IRO itself. The IRO is a possibility, the facts should be concrete claims about reality.
3. **Verify Facts**: Use the search tools to find whether these stated facts (not the IRO) are actually supported by evidence

CRITICAL: IRO STATEMENTS MUST USE MODAL OR HEDGED LANGUAGE
- IRO statements MUST include modal verbs (may, might, could, would, should) or hedging language (likely, potentially, possibly, appears to, seems to)
- IROs express POSSIBILITIES, UNCERTAINTIES, or POTENTIAL OUTCOMES - never state them as definitive facts
- Examples of CORRECT IRO statements:
  ✓ "The company may face water scarcity risks in its operations"
  ✓ "Climate change could impact supply chain resilience"
  ✓ "The organization might have opportunities to reduce carbon emissions"
  ✓ "Data privacy regulations could potentially affect business operations"
- Examples of INCORRECT IRO statements (too definitive):
  ✗ "The company faces water scarcity risks" (no modal language)
  ✗ "Climate change impacts supply chains" (stated as fact)
  ✗ "The organization has opportunities to reduce emissions" (too certain)

CRITICAL: SEPARATE IRO FROM FACTS
- IRO = the possibility/risk/opportunity being evaluated (from the question) - MUST use modal/hedging language
- Facts = concrete factual claims from the context that support evaluation - stated directly without hedging
- DO NOT include the IRO statement or the inference leading to the IRO statement in the list of facts
- Example:
  * IRO: "The company may face water scarcity risks" (possibility with modal verb "may")
  * Facts: "The company operates in water-stressed regions", "Water consumption increased by 20%" (concrete claims)
  * DO NOT include the risk/possibility itself as a fact

VERIFICATION STRATEGY:
1. **First**, search the document collection for each stated fact
2. **If facts are not found or partially found** in documents, use web_search as a fallback to verify missing facts
3. Clearly indicate which sources verified which facts (documents vs web)

VERIFICATION PRINCIPLES:

For **Negative Facts** (risks, incidents, violations, problems):
- Require explicit, strong evidence
- Do NOT assume problems exist without clear documentation
- Be fair and evidence-based
- Example: Context claims "data breach occurred" → needs explicit mention in documents, not just general security discussion

VERIFICATION LEVELS:
- 'yes': All stated facts are supported by documents OR web search (or both). Finding evidence in ANY source is good!
- 'partial': Only SOME of the stated facts are supported, but not all. 
- 'no': The stated facts are NOT supported by documents or web search, OR insufficient evidence found

IMPORTANT: If you find clear evidence for a fact in documents OR web, that fact is VERIFIED. Don't be overly strict!

OUTPUT FORMAT:
1. IRO Statement: One clear sentence describing the Impact/Risk/Opportunity (the POSSIBILITY). MUST use modal verbs (may, might, could, would, should) or hedging language (likely, potentially, possibly, appears to).
2. Stated Facts: List ONLY the factual claims from the context (bullet points or numbered) - EXCLUDE the IRO itself
3. Facts Verified: yes/no/partial
4. Evidence Description: What you found (quote/paraphrase) and how it relates to the stated facts. Clearly indicate for each fact whether it was verified by documents or web search.
5. Sources Consulted: List ALL sources used - format as "Document Name (page X)" for documents and "Web: [URL]" for web sources
6. New Question: Generate a survey question for materiality assessment (1-5 magnitude scale). First, infer the most relevant materiality dimension, then create a question tailored to that dimension. See detailed instructions below.
7. New Context: Reformulate the VERIFIED facts (only those with 'yes' or 'partial' verification) into 2-4 clear, factual sentences. Present only confirmed information without speculation. IMPORTANT: Do NOT include any source mentions, document names, page numbers, URLs, or citations - only the factual content.

Be precise, objective, and fair in your analysis."""

        # User query
        web_search_instruction = ""
        if has_web_search:
            web_search_instruction = "\n4. If any facts are not found in documents, use web_search to verify them as a fallback"
        
        # Get pattern enforcement instruction
        pattern_instruction = self._get_pattern_enforcement_instruction()
        
        query = f"""Please analyze the following IRO (Impact/Risk/Opportunity):

**QUESTION:**
{question}

**CONTEXT (containing stated facts):**
{context}

Please:
1. Extract the IRO statement from the question (1 sentence about the POSSIBILITY) - **MUST use modal/hedging language**
2. Identify ONLY the factual claims from the context - DO NOT include the IRO itself in the facts
3. Search the document collection first to verify whether these factual claims are supported{web_search_instruction}
4. Provide verification results with evidence and sources (clearly indicating document vs web sources)
5. Generate a NEW QUESTION using the dimension-based approach:
{pattern_instruction}
6. Generate a NEW CONTEXT: Reformulate ONLY the verified facts into 2-4 clear sentences (NO sources, citations, or document references - just the factual content)

CRITICAL REMINDERS: 
- IRO = the possibility/risk/opportunity (from question) - **MUST use modal verbs (may, might, could) or hedging language (likely, potentially, possibly)**
- Facts = concrete factual claims (from context) that support evaluation
- The IRO should NOT appear in the list of facts
- We verify the facts, NOT the IRO possibility
- Use web_search as fallback for facts not found in documents
- Clearly indicate sources: "Document Name (page X)" or "Web: [URL]"
- **IRO statements MUST be hedged/modal - expressing possibility, not certainty**
- New Question: **CRITICAL** - Infer the materiality dimension first, then use dimension-specific phrasing. DO NOT use "To what extent". Keep it GENERIC (not company-specific). Keep under 28 words. NO explanations, notes, or extra fields - ONLY the question. Ensure stylistic variation.
- New Context: Include ONLY verified facts (yes/partial), presented clearly. **NO source mentions, citations, document names, page numbers, or URLs - just the facts**

VERIFICATION APPROACH:
- Be thorough but NOT overly strict
- If you find clear evidence in documents OR web → the fact is VERIFIED
- Only mark as "partial" if SOME facts are verified but OTHERS are not
- Only mark as "no" if facts are clearly NOT supported by any source
- Don't penalize for using web vs documents - both are valid sources!"""

        try:
            # Use tool calling engine with structured output
            result = await self.engine.query(
                question=query,
                response_format=VerificationResult,
                system_prompt=system_prompt
            )
            
            # Clean all text fields to remove LLM artifacts and sanitize for YAML/Excel output
            result.iro_statement = clean_text(result.iro_statement)
            result.stated_facts = clean_text(result.stated_facts)
            result.facts_verified = clean_text(result.facts_verified)
            result.evidence_description = clean_text(result.evidence_description)
            result.sources_consulted = clean_text(result.sources_consulted)
            result.new_question = clean_text(result.new_question)
            result.new_context = clean_text(result.new_context)
            
            # Track the pattern used for the new question to enforce variation
            self._track_question_pattern(result.new_question)
            
            logger.info(f"Verification complete: {result.facts_verified}")
            logger.info(f"New question pattern: {result.new_question[:50]}...")
            return result
            
        except Exception as e:
            logger.error(f"Error during verification: {e}")
            # Return a default result indicating an error
            return VerificationResult(
                iro_statement="Error",
                stated_facts="N/A",
                facts_verified="error",
                evidence_description=f"Error occurred during verification: {str(e)}",
                sources_consulted="N/A",
                new_question="N/A",
                new_context="N/A"
            )
    
    def _load_checkpoints(self, checkpoint_file: Path) -> Dict[int, IROCheckpoint]:
        """Load checkpoints from YAML file."""
        if not checkpoint_file.exists():
            return {}
        
        try:
            with open(checkpoint_file, 'r') as f:
                data = yaml.safe_load(f) or {}
            
            checkpoints = {}
            for row_idx_str, checkpoint_data in data.items():
                row_idx = int(row_idx_str)
                # Convert result dict to VerificationResult if present
                if checkpoint_data.get('result'):
                    result_data = checkpoint_data['result']
                    # Clean text fields when loading from checkpoint
                    if 'iro_statement' in result_data:
                        result_data['iro_statement'] = clean_text(result_data['iro_statement'])
                    if 'stated_facts' in result_data:
                        result_data['stated_facts'] = clean_text(result_data['stated_facts'])
                    if 'facts_verified' in result_data:
                        result_data['facts_verified'] = clean_text(result_data['facts_verified'])
                    if 'evidence_description' in result_data:
                        result_data['evidence_description'] = clean_text(result_data['evidence_description'])
                    if 'sources_consulted' in result_data:
                        result_data['sources_consulted'] = clean_text(result_data['sources_consulted'])
                    if 'new_question' in result_data:
                        result_data['new_question'] = clean_text(result_data['new_question'])
                    if 'new_context' in result_data:
                        result_data['new_context'] = clean_text(result_data['new_context'])
                    checkpoint_data['result'] = VerificationResult(**result_data)
                checkpoints[row_idx] = IROCheckpoint(**checkpoint_data)
            
            logger.info(f"Loaded {len(checkpoints)} checkpoints from {checkpoint_file}")
            return checkpoints
        except Exception as e:
            logger.error(f"Error loading checkpoints: {e}")
            return {}
    
    def _save_checkpoints(self, checkpoint_file: Path, checkpoints: Dict[int, IROCheckpoint]):
        """Save checkpoints to YAML file."""
        try:
            # Convert to dict for YAML serialization
            data = {}
            for row_idx, checkpoint in checkpoints.items():
                checkpoint_dict = checkpoint.model_dump()
                # Convert result to dict if present
                if checkpoint_dict.get('result'):
                    checkpoint_dict['result'] = checkpoint_dict['result']
                data[str(row_idx)] = checkpoint_dict
            
            with open(checkpoint_file, 'w') as f:
                yaml.safe_dump(data, f, default_flow_style=False, sort_keys=False)
            
            logger.info(f"Saved {len(checkpoints)} checkpoints to {checkpoint_file}")
        except Exception as e:
            logger.error(f"Error saving checkpoints: {e}")
    
    async def _process_single_row(
        self,
        row_idx: int,
        question: str,
        context: str,
        semaphore: asyncio.Semaphore,
        progress: Progress,
        task: TaskID,
        retry_count: int = 0
    ) -> Tuple[int, VerificationResult, bool]:
        """
        Process a single row with rate limiting.
        
        Args:
            row_idx: Row index in the spreadsheet
            question: Question that evaluates the IRO
            context: Context containing stated facts
            semaphore: Asyncio semaphore for concurrency control
            progress: Progress bar instance
            task: Task ID for progress updates
            retry_count: Current retry attempt number
            
        Returns:
            Tuple of (row_idx, VerificationResult, success: bool)
        """
        async with semaphore:
            # Log which row is being processed
            retry_suffix = f" (retry {retry_count}/{MAX_RETRIES})" if retry_count > 0 else ""
            question_preview = question[:80] + "..." if len(question) > 80 else question
            self.console.print(f"[yellow]Row {row_idx}{retry_suffix}:[/yellow] Processing: {question_preview}")
            logger.info(f"Row {row_idx}: Starting IRO verification{retry_suffix}")
            
            success = False
            try:
                result = await self.verify_iro(question, context)
                self.console.print(f"[green]Row {row_idx}:[/green] Complete - Facts Verified: {result.facts_verified}")
                logger.info(f"Row {row_idx}: Verification complete - {result.facts_verified}")
                success = True
            except Exception as e:
                self.console.print(f"[red]Row {row_idx}:[/red] Error - {str(e)}")
                logger.error(f"Row {row_idx}: Error during verification - {str(e)}")
                result = VerificationResult(
                    iro_statement="Error",
                    stated_facts="N/A",
                    facts_verified="error",
                    evidence_description=f"Error occurred: {str(e)}",
                    sources_consulted="N/A"
                )
            
            progress.update(task, advance=1)
            return (row_idx, result, success)
    
    async def verify_iros_from_excel(
        self,
        input_file: Path,
        output_file: Path,
        question_column: str = "questions",
        context_column: str = "contexts",
        max_concurrent: int = MAX_CONCURRENT_VERIFICATIONS,
        batch_size: int = BATCH_SIZE,
        limit: Optional[int] = None
    ) -> None:
        """
        Read IROs from an Excel file and verify each one in parallel with checkpointing and retry logic.
        
        Args:
            input_file: Path to input XLSX file
            output_file: Path to output XLSX file
            question_column: Name of the column containing IRO questions
            context_column: Name of the column containing contexts with stated facts
            max_concurrent: Maximum number of concurrent verifications (default: 5)
            batch_size: Number of IROs to process before saving checkpoint (default: 10)
            limit: Optional limit on number of IROs to process (for testing)
        """
        # Setup checkpoint file
        checkpoint_file = input_file.parent / f".checkpoint_{input_file.stem}.yaml"
        
        # Load existing checkpoints if any
        checkpoints = self._load_checkpoints(checkpoint_file)
        
        if checkpoints:
            self.console.print(f"[cyan]Found existing checkpoint with {len(checkpoints)} IROs[/cyan]")
            completed_count = sum(1 for cp in checkpoints.values() if cp.status == 'completed')
            self.console.print(f"[dim]  - Completed: {completed_count}[/dim]")
            self.console.print(f"[dim]  - Will resume from checkpoint[/dim]\n")
            
            # Restore pattern tracking from completed checkpoints to maintain variation
            completed_checkpoints = sorted(
                [cp for cp in checkpoints.values() if cp.status == 'completed' and cp.result],
                key=lambda x: x.row_idx
            )
            for cp in completed_checkpoints[-5:]:  # Last 5 completed
                if cp.result and cp.result.new_question:
                    self._track_question_pattern(cp.result.new_question)
            
            if self.recent_question_patterns:
                logger.info(f"Restored {len(self.recent_question_patterns)} recent question patterns from checkpoint")
                self.console.print(f"[dim]  - Restored pattern variation state[/dim]")
        
        # Load the workbook
        logger.info(f"Loading input file: {input_file}")
        wb = load_workbook(input_file)
        ws = wb.active
        
        # Find the question and context columns
        headers = [cell.value for cell in ws[1]]
        
        # Filter out None values for display purposes
        valid_headers = [h for h in headers if h is not None]
        
        # Helper function to find column with flexible naming (singular/plural)
        def find_column(preferred_name: str, alternatives: List[str]) -> Tuple[str, int]:
            """
            Find a column by trying multiple name variants.
            Returns (actual_column_name, column_index_1based)
            """
            all_options = [preferred_name] + alternatives
            for option in all_options:
                if option in headers:
                    return option, headers.index(option) + 1
            
            # Not found - raise error
            raise ValueError(
                f"Column '{preferred_name}' (or alternatives: {', '.join(alternatives)}) "
                f"not found in spreadsheet. Available columns: {', '.join(valid_headers)}"
            )
        
        # Try to find question column (support: question, questions, description, descriptions)
        question_alternatives = []
        if question_column == "questions":
            question_alternatives = ["question", "description", "descriptions"]
        elif question_column == "question":
            question_alternatives = ["questions", "description", "descriptions"]
        else:
            question_alternatives = ["question", "questions", "description", "descriptions"]
        
        actual_question_col, question_col_idx = find_column(question_column, question_alternatives)
        
        # Try to find context column (support: context, contexts)
        context_alternatives = []
        if context_column == "contexts":
            context_alternatives = ["context"]
        elif context_column == "context":
            context_alternatives = ["contexts"]
        else:
            context_alternatives = ["context", "contexts"]
        
        actual_context_col, context_col_idx = find_column(context_column, context_alternatives)
        
        logger.info(f"Using columns: '{actual_question_col}' and '{actual_context_col}'")
        self.console.print(f"[cyan]📋 Using columns:[/cyan] '{actual_question_col}' and '{actual_context_col}'")
        
        # Collect all rows (create checkpoints for new ones)
        for row_idx in range(2, ws.max_row + 1):  # Start from 2 (skip header)
            question = ws.cell(row=row_idx, column=question_col_idx).value
            context = ws.cell(row=row_idx, column=context_col_idx).value
            
            # Skip empty rows
            if not question or not str(question).strip() or not context or not str(context).strip():
                continue
            
            # Create checkpoint if doesn't exist
            if row_idx not in checkpoints:
                checkpoints[row_idx] = IROCheckpoint(
                    row_idx=row_idx,
                    question=str(question),
                    context=str(context),
                    status='pending'
                )
        
        # Apply limit if specified (for testing)
        all_row_indices = sorted(checkpoints.keys())
        if limit and limit > 0:
            original_count = len(all_row_indices)
            all_row_indices = all_row_indices[:limit]
            self.console.print(f"[yellow]Testing mode: Processing first {len(all_row_indices)} of {original_count} IROs[/yellow]")
            logger.info(f"Testing mode: Limiting to first {len(all_row_indices)} IROs (out of {original_count})")
        
        # Identify pending IROs (including errors that need retry)
        pending_indices = [
            idx for idx in all_row_indices 
            if checkpoints[idx].status != 'completed'
        ]
        
        completed_count = len(all_row_indices) - len(pending_indices)
        total_count = len(all_row_indices)
        
        logger.info(f"Processing {total_count} IROs ({completed_count} already completed, {len(pending_indices)} pending)")
        self.console.print(f"\n[bold cyan]Verification Status:[/bold cyan]")
        self.console.print(f"  Total IROs: {total_count}")
        self.console.print(f"  Completed: {completed_count}")
        self.console.print(f"  Pending: {len(pending_indices)}")
        self.console.print(f"  Batch size: {batch_size}")
        self.console.print(f"  Max concurrent: {max_concurrent}\n")
        
        if not pending_indices:
            self.console.print("[green]All IROs already completed! Generating Excel file...[/green]\n")
        else:
        # Create semaphore for rate limiting
            semaphore = asyncio.Semaphore(max_concurrent)
        
            # Process in batches
            total_processed = 0
            
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TextColumn("({task.completed}/{task.total})"),
            TimeElapsedColumn(),
            console=self.console
        ) as progress:
                task = progress.add_task("[cyan]Verifying IROs...", total=len(pending_indices))
                
                # Process in batches
                for batch_start in range(0, len(pending_indices), batch_size):
                    batch_indices = pending_indices[batch_start:batch_start + batch_size]
                    batch_num = (batch_start // batch_size) + 1
                    total_batches = (len(pending_indices) + batch_size - 1) // batch_size
                    
                    self.console.print(f"\n[cyan]═══ Batch {batch_num}/{total_batches} ({len(batch_indices)} IROs) ═══[/cyan]")
                    
                    # Create tasks for this batch
                    tasks = []
                    for idx in batch_indices:
                        cp = checkpoints[idx]
                        tasks.append(
                            self._process_single_row(
                                cp.row_idx, cp.question, cp.context,
                                semaphore, progress, task, cp.retry_count
                            )
                        )
                    
                    # Execute batch
                    batch_results = await asyncio.gather(*tasks)
                    
                    # Update checkpoints with results
                    for row_idx, result, success in batch_results:
                        cp = checkpoints[row_idx]
                        cp.result = result
                        cp.last_updated = datetime.now().isoformat()
                        
                        if success:
                            cp.status = 'completed'
                            cp.retry_count = 0
                        else:
                            cp.retry_count += 1
                            if cp.retry_count >= MAX_RETRIES:
                                cp.status = 'error'
                                cp.error_message = f"Failed after {MAX_RETRIES} retries"
                                self.console.print(f"[red]Row {row_idx}: Max retries reached[/red]")
                            else:
                                cp.status = 'pending'  # Will retry
                        
                        total_processed += 1
                    
                    # Save checkpoint after batch
                    self._save_checkpoints(checkpoint_file, checkpoints)
                    self.console.print(f"[dim]✓ Checkpoint saved ({total_processed}/{len(pending_indices)} processed)[/dim]")
            
            # Retry failed IROs (those with status='pending' and retry_count > 0)
                retry_indices = [
                idx for idx in all_row_indices
                if checkpoints[idx].status == 'pending' and checkpoints[idx].retry_count > 0
                ]
            
                if retry_indices:
                    self.console.print(f"\n[yellow]Retrying {len(retry_indices)} failed IROs...[/yellow]")
                    # Recursively retry by updating pending_indices and continuing
                    # For simplicity, we'll just log them here and they'll be picked up on next run
                    for idx in retry_indices:
                        cp = checkpoints[idx]
                        self.console.print(f"  Row {idx}: Retry {cp.retry_count}/{MAX_RETRIES}")
        
        # Generate Excel file from checkpoints
        self.console.print("\n[yellow]Generating Excel file from results...[/yellow]")
        
        # Add headers if needed
        new_headers = [
            "IRO Statement",
            "Stated Facts",
            "Facts Verified (yes/no/partial)",
            "Evidence Description",
            "Sources Consulted",
            "New Question",
            "New Context"
        ]
        for i, header in enumerate(new_headers, start=1):
            cell = ws.cell(row=1, column=len(headers) + i)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Write results
        for row_idx, cp in checkpoints.items():
            if cp.result:
                ws.cell(row=row_idx, column=len(headers) + 1).value = cp.result.iro_statement
                ws.cell(row=row_idx, column=len(headers) + 2).value = cp.result.stated_facts
                ws.cell(row=row_idx, column=len(headers) + 3).value = cp.result.facts_verified
                ws.cell(row=row_idx, column=len(headers) + 4).value = cp.result.evidence_description
                ws.cell(row=row_idx, column=len(headers) + 5).value = cp.result.sources_consulted
                ws.cell(row=row_idx, column=len(headers) + 6).value = cp.result.new_question
                ws.cell(row=row_idx, column=len(headers) + 7).value = cp.result.new_context
            else:
                # Empty or skipped row
                ws.cell(row=row_idx, column=len(headers) + 1).value = "N/A"
                ws.cell(row=row_idx, column=len(headers) + 2).value = "N/A"
                ws.cell(row=row_idx, column=len(headers) + 3).value = "N/A"
                ws.cell(row=row_idx, column=len(headers) + 4).value = cp.error_message or "Not processed"
                ws.cell(row=row_idx, column=len(headers) + 5).value = "N/A"
                ws.cell(row=row_idx, column=len(headers) + 6).value = "N/A"
                ws.cell(row=row_idx, column=len(headers) + 7).value = "N/A"
        
        # Save Excel file
        logger.info(f"Saving results to: {output_file}")
        wb.save(output_file)
        
        # Final summary
        final_completed = sum(1 for cp in checkpoints.values() if cp.status == 'completed')
        final_errors = sum(1 for cp in checkpoints.values() if cp.status == 'error')
        
        self.console.print(f"\n✅ [bold green]Verification complete![/bold green]")
        self.console.print(f"[dim]Results saved to: {output_file}[/dim]")
        self.console.print(f"\n[bold]Summary:[/bold]")
        self.console.print(f"  Total IROs: {total_count}")
        self.console.print(f"  ✓ Completed: {final_completed}")
        if final_errors > 0:
            self.console.print(f"  ✗ Errors: {final_errors}")
        
        # Always preserve checkpoint file with intermediate results
        self.console.print(f"\n[cyan]Checkpoint saved to: {checkpoint_file}[/cyan]")
        if final_errors == 0 and final_completed == total_count:
            self.console.print(f"[dim]All IROs completed successfully - checkpoint contains full results[/dim]")
        else:
            self.console.print(f"[dim]Run again to retry failed IROs[/dim]")
        
        logger.info("Verification complete!")


async def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Verify IROs (Impacts, Risks, Opportunities) in an Excel file against a document collection using RAG"
    )
    parser.add_argument(
        "--input",
        type=str,
        required=True,
        help="Path to input XLSX file containing IROs to verify"
    )
    parser.add_argument(
        "--collection",
        type=str,
        required=True,
        help="Name of the document collection to search against"
    )
    parser.add_argument(
        "--output",
        type=str,
        help="Path to output XLSX file (default: input file with '_verified' suffix)"
    )
    parser.add_argument(
        "--question-column",
        type=str,
        default="questions",
        help="Name of the column containing IRO questions (default: questions). Also accepts: question, description, descriptions"
    )
    parser.add_argument(
        "--context-column",
        type=str,
        default="contexts",
        help="Name of the column containing contexts with stated facts (default: contexts). Also accepts: context"
    )
    parser.add_argument(
        "--model",
        type=str,
        default=DEFAULT_MODEL,
        help=f"OpenAI model to use (default: {DEFAULT_MODEL})"
    )
    parser.add_argument(
        "--concurrent",
        type=int,
        default=MAX_CONCURRENT_VERIFICATIONS,
        help=f"Maximum number of concurrent verifications (default: {MAX_CONCURRENT_VERIFICATIONS})"
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=BATCH_SIZE,
        help=f"Number of IROs to process before saving checkpoint (default: {BATCH_SIZE})"
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Limit processing to first N IROs (for testing)"
    )
    parser.add_argument(
        "--no-rerank",
        action="store_true",
        help="Disable Cohere reranking to avoid rate limits (faster, less accurate)"
    )
    
    args = parser.parse_args()
    
    # Temporarily disable reranking if requested
    if args.no_rerank:
        os.environ["USE_COHERE_RERANK"] = "false"
        logger.info("Cohere reranking disabled for this run")
    
    # Validate input file
    input_file = Path(args.input)
    if not input_file.exists():
        print(f"Error: Input file not found: {input_file}")
        sys.exit(1)
    
    # Determine output file
    if args.output:
        output_file = Path(args.output)
    else:
        # Generate output filename with 'results_' prefix
        output_file = input_file.parent / f"results_{input_file.name}"
    
    # Create verifier and process
    verifier = StatementVerifier(
        collection=args.collection,
        model=args.model
    )
    
    await verifier.verify_iros_from_excel(
        input_file=input_file,
        output_file=output_file,
        question_column=args.question_column,
        context_column=args.context_column,
        max_concurrent=args.concurrent,
        batch_size=args.batch_size,
        limit=args.limit
    )


if __name__ == "__main__":
    asyncio.run(main())

